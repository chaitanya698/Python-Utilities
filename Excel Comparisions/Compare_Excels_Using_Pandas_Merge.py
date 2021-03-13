
import pandas as pd
import pandas
import openpyxl as opxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import xlsxwriter as xw
import xlrd
import os
import datetime
import numpy as np
from pathlib import Path
try:
    from StringIO import StringIO
except ImportError:
    from io import StringIO

def dataframe_difference(file1, file2, sheet_name, fname, which=None):

    df1 = pd.read_excel(file1, sheet_name=sheet_name).fillna('')
    df2 = pd.read_excel(file2, sheet_name=sheet_name).fillna('')

    which = None
    
# Here we implement the DataFrame Merge to get only the different rows from both the files.

    comparison_df = df1.merge(df2, indicator=True, sort=True, how='outer')

    if which is None:
        diff_df = comparison_df[comparison_df['_merge'] != 'both']
    else:
        diff_df = comparison_df[comparison_df['_merge'] == which]

# This is for getting the merge column to the begining

    cols = list(diff_df.columns)
    cols = [cols[-1]] + cols[:-1]
    diff_df = diff_df[cols]

    diff_df.sort_values(by='_merge', inplace=True)
    # print(diff_df)
    
    df_OLD = diff_df.loc[diff_df['_merge'] == 'left_only']
    df_NEW = diff_df.loc[diff_df['_merge'] == 'right_only']

    df_NEW.sort_values(by='_merge', inplace=True)
    df_OLD.sort_values(by='_merge', inplace=True)
    
    df_NEW.drop(['_merge'], axis=1, inplace=True)
    df_OLD.drop(['_merge'], axis=1, inplace=True)

    cols_OLD = df_OLD.columns
    cols_NEW = df_NEW.columns
    sharedCols = list(set(cols_OLD).intersection(cols_NEW))

# Here We Perform Diffences and capture the new and old values

    dfDiff = df_OLD.copy()

    for row in range(dfDiff.shape[0]):
        for col in range(dfDiff.shape[1]):
            try:
                value_OLD = df_OLD.iloc[row, col]
                value_NEW = df_NEW.iloc[row, col]
                # print(value_NEW)
                if value_OLD == value_NEW:
                    dfDiff.iloc[row, col] = df_NEW.iloc[row, col]
                else:
                    if value_NEW == '':
                        value_NEW = 'Null'
                    if value_OLD == '':
                        value_OLD = 'Null'
                    dfDiff.iloc[row,col] = ('Pre:{}→ Post:{}').format(value_OLD, value_NEW)
                    Sheets_changed.append('{}'.format(sheet_name))
                    discrepancies.append("Data Miss Match")
                    rows_changed.append('Row: {}'.format(col))
                    cols_changed.append('Column: {}'.format(col))
            except Exception as e:
                dfDiff.iloc[row, col] = ('{}→{}').format(value_OLD, 'NaN')
    
# Here we are finding the Missing and Added rows

    missing_rows = set(df1.index) - set(df2.index)

    added_rows = set(df2.index) - set(df1.index)

    for row in missing_rows:
        Sheets_changed.append('{}'.format(sheet_name))
        discrepancies.append('Missing Data')
        rows_changed.append('Row: {}'.format(row))
    
    result2 = df1.merge(df2, how = 'outer' ,indicator=True).loc[lambda x : x['_merge']=='right_only']
    
    cols = list(result2.columns)
    cols = [cols[-1]] + cols[:-1]
    result2 = result2[cols]

    result2.replace('right_only', 'Added_Rows_Post', inplace=True)
    
    for row in added_rows:
        Sheets_changed.append('{}'.format(sheet_name))
        discrepancies.append('New Data')
        rows_changed.append('Row: {}'.format(row))
        cols_changed.append('Column: {}'.format(''))
    
    book = load_workbook(fname)
    writer = pd.ExcelWriter(fname, engine='openpyxl')
    writer.book = book
    result2.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
    dfDiff.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
    # worksheet = book.active
    # worksheet = writer.sheets[sheet_name]
    writer.save()
    dfDiff.reset_index(inplace=False, drop=True)
    print(sheet_name + ' Compared.!\n')
  
def write_differences(fname,Sheets_changed,rows_changed,discrepancies,cols_changed,Status,rep_name):
    
    if(len(Sheets_changed)>0):
        Status.append("Failed")
        fail_count.append(rep_name)
    else:
        Status.append("Passed")
        pass_count.append(rep_name)
    book = load_workbook(fname)
    writer = pd.ExcelWriter(fname, engine='openpyxl')
    writer.book = book
    std = book['Sheet']
    book.remove(std) 
    rep_name = StringIO(rep_name)
    report = pd.DataFrame(rep_name,columns=['Report Name'])
    status = pd.DataFrame(Status,columns=['Status'])
#     Excep = pd.DataFrame(Exceptions,columns=['Exception Occured'])
    report.to_excel(writer, sheet_name='High Level Summary',startrow=0,header=True,index=False)
    status.to_excel(writer, sheet_name='High Level Summary',startcol=1,header=True,index=False)
#     Excep.to_excel(writer, sheet_name='High Level Summary',startcol=2,header=True,index=False)
    
    Sheets = pd.DataFrame(Sheets_changed, columns=['Sheet Name'])
    discrepancies_dat = pd.DataFrame(discrepancies, columns=['Type Of Discrepancy'])
    rows_chd = pd.DataFrame(rows_changed, columns=['Row Number'])
    clos_chd = pd.DataFrame(cols_changed, columns=['Column Number'])

    Sheets.to_excel(writer, sheet_name='Detailed Summary', header=True, index=False)
    discrepancies_dat.to_excel(writer, sheet_name='Detailed Summary', startcol=1,header=True, index=False)
    rows_chd.to_excel(writer, sheet_name='Detailed Summary',startcol=2,header=True,index=False)
    clos_chd.to_excel(writer,sheet_name='Detailed Summary',startcol=3,header=True,index=False)
    
    sheets=book._sheets
    from_loc = None
    to_loc = None
    # if no from_loc given, assume last sheet
    if from_loc is None:
        from_loc = len(sheets) - 1

    #if no to_loc given, assume first
    if to_loc is None:
        to_loc = 0

    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)
    
    worksheet = writer.sheets['High Level Summary']
    worksheet = book._sheets
    from_loc = None
    to_loc = None
    if from_loc is None:
        from_loc = len(worksheet) - 1
    if to_loc is None:
        to_loc = 0
    sheet = worksheet.pop(from_loc)
    worksheet.insert(to_loc, sheet)

    writer.save()
    print('\n Differences Wiritten.!\n')
    
def remove_blank_sheets(fname,Sheets_changed):
    try:
        sheets=[]
        workbook=opxl.load_workbook(fname)
        sheets = list(set(Sheets_changed))
        for sheet in workbook.sheetnames:
            if ((sheet not in sheets) and sheet !='Detailed Summary' and sheet !='High Level Summary'):
                workbook.remove(workbook[sheet])
        workbook.save(fname)
    except Exception as e:
        print("No Sheets")
        
def add_formatting(fname):

    wbRD = xlrd.open_workbook(fname)
    sheets = wbRD.sheets()
    workbook = xw.Workbook(fname)
    writer = pd.ExcelWriter(fname, engine='xlsxwriter')

    for sheet in sheets:
        sheet1 = workbook.add_worksheet(sheet.name)
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                sheet1.write(row, col, sheet.cell(row, col).value)

        sheet1.set_column('A1:ZZ1000', 25)
        sheet1.hide_gridlines(2)
        sheet1.set_default_row(15)

    # define formats
    
        highlight_fmt = workbook.add_format({
            'bold': True,
            'font_name': 'Gill Sans MT',
            'font_color': '#25679B',
            'bg_color': '#4AF093'
        })
        new_fmt = workbook.add_format({
            'font_name': 'Gill Sans MT',
            'font_color': '#4A4F47',
            'bold': False
        })
        header_fmt = workbook.add_format({
            'font': 'Gill Sans MT',
            'font_color': '#4A4F47',
            'bg_color': '#99ccff',
            'bold': True
        })
        top_header_fmt = workbook.add_format({
            'font': 'Gill Sans MT',
            'font_color': '#4A4F47',
            'bg_color': '#4AA7F0',
            'bold': True
        })
        
        missing_row_fmt = workbook.add_format({
            'font': 'Gill Sans MT',
            'bold': True,
            'font_color': '#f2f2f2',
            'bg_color': '#DC143C'
        })
        
        added_row_fmt = workbook.add_format({
            'font': 'Gill Sans MT',
            'bold': True,
            'font_color': '#25679B',
            'bg_color': 'e6e6ff'
        })

        
    # highlight the missing rows
    
        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': '→NaN',
                'format': missing_row_fmt
            })
        
        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Added_Rows',
                'format': added_row_fmt
            })
        
    # highlight unchanged cells
    
        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'not containing',
                'value': '→',
                'format': new_fmt
            })
        
    # highlight changed cells
    
        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': '→',
                'format': highlight_fmt
            })

        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Sheet Name',
                'format': header_fmt
            })

        sheet1.conditional_format(
        'A1:ZZ1000', {
            'type': 'text',
            'criteria': 'containing',
            'value': 'Type Of Discrepancy',
            'format': header_fmt
        })

        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Row Number',
                'format': header_fmt
            })

        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Column Number',
                'format': header_fmt
            })

        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Report Name',
                'format': header_fmt
            })

        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Status',
                'format': header_fmt
            })
        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Exception Occured',
                'format': header_fmt
            })
        sheet1.conditional_format(
            'A1:ZZ1000', {
                'type': 'text',
                'criteria': 'containing',
                'value': 'Differences Summary',
                'format': top_header_fmt
            })

    workbook.close()
    
if __name__ == '__main__':
    
    file1= input("Enter the file1 name:")
    file2= input("Enter the file2 name:")
    
    file1 = Path(file1)
    file2 = Path(file2)
    fname = 'Differences_{} vs {}.xlsx'.format(file1.stem, file2.stem)
    old_rows = []
    appended_rows = []
    rows = []
    rows_changed = []
    Sheets_changed = []
    cols_changed = []
    discrepancies = []
    Status=[]
    fail_count=[]
    pass_count=[]
    rep_name= fname
    print(rep_name)

    df_missing = pd.DataFrame()
    df_added = pd.DataFrame()
    wb = opxl.Workbook()
    wb.save(fname)
    xls3 = pd.ExcelFile(file1)
    xls4 = pd.ExcelFile(fname)

    for sheet_name in xls3.sheet_names:
        dataframe_difference(file1, file2, sheet_name, fname, which=None)

    write_differences(fname,Sheets_changed,rows_changed,discrepancies,cols_changed,Status,rep_name)    
    add_formatting(fname)
    remove_blank_sheets(fname,Sheets_changed)

