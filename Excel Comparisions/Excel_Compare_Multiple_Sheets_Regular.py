#!/usr/bin/env python
# coding: utf-8

# In[ ]:


##############################
# Excel Comparator          #
# Author: Chaitanya Annangi #
##############################

import pandas as pd
from pathlib import Path
import openpyxl as opxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import xlsxwriter as xw
import xlrd
import os

def excel_diff(path_OLD, path_NEW,sheet_name):

    df_OLD = pd.read_excel(path_OLD,sheet_name=sheet_name,header=1).fillna('')
    df_NEW = pd.read_excel(path_NEW,sheet_name=sheet_name,header=1).fillna('')
    
    cols_OLD = df_OLD.columns
    cols_NEW = df_NEW.columns
    sharedCols = list(set(cols_OLD).intersection(cols_NEW))

    # Perform Diffences
    dfDiff = df_NEW.copy()
    try:
        for row in dfDiff.index:
            if (row in df_NEW.index):
                for col in sharedCols:
                    value_OLD = df_OLD.loc[row,col]
                    value_NEW = df_NEW.loc[row,col]
                    if value_OLD==value_NEW:
                        dfDiff.loc[row,col] = df_NEW.loc[row,col]
                    else:
                        if value_NEW == '':
                            value_NEW='Null'
                        if value_OLD == '':
                            value_OLD='Null'
                        dfDiff.loc[row,col] = ('Expected:{}→ Actual:{}').format(value_OLD,value_NEW) 
                        Sheets_changed.append(' {}'.format(sheet_name))
                        rows_changed.append('Row: {}'.format(row))
                        cols_changed.append('Column: {}'.format(col.strip('Unnamed:'))) 
                        
    except Exception as e:
        print(e)

    # Here we Identify the Missing Rows and Newly Added Rows

    missing_rows = set(df_OLD.index) - set(df_NEW.index)
    added_rows = set(df_NEW.index) - set(df_OLD.index)
    for row in missing_rows:
        old_rows.append('Sheet_Name:{}, Row: {}, Column: {}'.format(sheet_name,row,col.strip('Unnamed:')))
    for row in added_rows:
        appended_rows.append('Sheet_Name:{}, Row: {}, Column: {}'.format(sheet_name,row,col.strip('Unnamed:')))
    
    book = load_workbook(fname)
    writer = pd.ExcelWriter(fname, engine='openpyxl') 
    writer.book = book
    dfDiff.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
    worksheet = book.active
    worksheet = writer.sheets[sheet_name]
    writer.save()
    dfDiff.reset_index(inplace=False,drop=True)
    print(sheet_name+' Compared.!\n')

    
def write_differences(fname):
    
    book = load_workbook(fname)
    writer = pd.ExcelWriter(fname, engine='openpyxl') 
    writer.book = book
    std=book['Sheet']
    book.remove(std)

    Sheets=pd.DataFrame(Sheets_changed,columns=['Sheet Name'])
    rows_chd=pd.DataFrame(rows_changed,columns=['Row Number'])
    clos_chd=pd.DataFrame(cols_changed,columns=['Column Number'])
    drop=pd.DataFrame(old_rows,columns=['Dropped Rows'])
    add=pd.DataFrame(appended_rows,columns=['Added Rows'])
    
    Sheets.to_excel(writer, sheet_name='Differences',header=True, index=False)
    rows_chd.to_excel(writer, sheet_name='Differences',startcol=1,header=True, index=False)
    clos_chd.to_excel(writer, sheet_name='Differences',startcol=2,header=True, index=False)
    add.to_excel(writer, sheet_name='Differences',startcol=3,header=True, index=False)
    drop.to_excel(writer, sheet_name='Differences',startcol=4,header=True, index=False)
    
    
    worksheet = writer.sheets['Differences']
    worksheet=book._sheets
    from_loc=None
    to_loc=None
    if from_loc is None:
        from_loc = len(worksheet) - 1
    if to_loc is None:
        to_loc = 0
    sheet = worksheet.pop(from_loc)
    worksheet.insert(to_loc, sheet)

    writer.save()
    print('\n Differences Wiritten.!\n')

def add_formatting(fname):

    wbRD = xlrd.open_workbook(fname)
    sheets = wbRD.sheets()
    workbook = xw.Workbook(fname)
    writer = pd.ExcelWriter(fname, engine='xlsxwriter')
    
    for sheet in sheets: 
        sheet1 = workbook.add_worksheet(sheet.name)
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                sheet1.write(row, col, sheet.cell(row, col).value)

        sheet1.set_column('A1:ZZ1000',25)
        sheet1.hide_gridlines(2)
        sheet1.set_default_row(15)
    
    # define formats
        highlight_fmt = workbook.add_format({'bold': True,'font_color': '#FF0000', 'bg_color':'#FFFF00'})
        new_fmt = workbook.add_format({'font_color': '#808080','bold':False})

    # highlight unchanged cells
        sheet1.conditional_format('A1:ZZ1000', { 'type': 'text',
                                                    'criteria': 'not containing',
                                                    'value':'→',
                                                    'format':new_fmt })
    # highlight changed cells
        sheet1.conditional_format('A1:ZZ1000', { 'type': 'text',
                                                    'criteria': 'containing',
                                                    'value':'→',
                                                    'format':highlight_fmt})

    workbook.close()

def main():
    
    xls1= pd.ExcelFile(path_OLD)
    xls2= pd.ExcelFile(path_NEW)
    
    if(xls1.sheet_names != xls2.sheet_names):
        print("Excel files are of not same shape.!")
    else:
        print("\nExcel files are of same shape, Starting the Comparision.!\n")
        print("No.of Pages in Excel Sheets:",len(xls1.sheet_names))
        print("\n")
        for sheet_name in xls2.sheet_names:
            excel_diff(path_OLD, path_NEW,sheet_name)  
        write_differences(fname)

    xls3 = pd.ExcelFile(fname)

    print("\nFormatting the Excel File.!\n")

    for sheet_name in xls3.sheet_names:
        add_formatting(fname)

    print("\nOutput File is generated.!\n")

if __name__ == '__main__':

    rows_changed=[]
    Sheets_changed=[]
    cols_changed=[]
    old_rows=[]
    appended_rows=[]
    file1= input("Enter the file1 name:")
    file2= input("Enter the file2 name:")
    path_OLD = Path(file1)
    path_NEW = Path(file2)

    fname = ".\\Differences.xlsx"
    wb = opxl.Workbook()
    wb.save(fname)    
    main()


# In[ ]:





# In[ ]:




